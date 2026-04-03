import io
import re
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MASTERDATA_PATH = Path(__file__).parent / "Masterdata.xlsx"

HOJA_ABM = "ABM ART. DEPURADO 23.02.2025"
HOJA_RESUMEN = "RESUMEN"
COL_CODIGO_DARNEL = 0
COL_CANT_EMPAQUE = 17

ZONE_BG = {
    1: "1F4E79",
    2: "2E75B6",
    3: "0070C0",
    4: "00B0F0",
    5: "9DC3E6",
    6: "BDD7EE",
}
ZONE_FG = {
    1: "FFFFFF",
    2: "FFFFFF",
    3: "FFFFFF",
    4: "1F3864",
    5: "1F3864",
    6: "1F3864",
}
HEADERS_D = [
    "Zona",
    "Código Darnel",
    "Código Pilarica",
    "Descripción",
    "Cant x Un Empaque",
    "Uds/Pallet",
    "Pallets",
]
COL_W_D = [8, 18, 17, 38, 18, 13, 13]

COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ZONA_BG = "2E75B6"
COLOR_ZONA_FG = "FFFFFF"
COLOR_TOTAL_BG = "D6E4F0"
COLOR_TOTAL_FG = "1F4E79"
COLOR_GTOTAL_BG = "1F4E79"
COLOR_GTOTAL_FG = "FFFFFF"
COLOR_ROW_ODD = "FFFFFF"
COLOR_ROW_EVEN = "EBF3FB"

COL_HEADERS_Z = [
    "Zona",
    "Código Artículo",
    "Descripción",
    "Cliente",
    "Cantidad",
    "Uds/Pallet",
    "Pallets",
]
COL_WIDTHS_Z = [8, 18, 35, 30, 12, 12, 10]

_thin_d = Side(style="thin", color="B0C4DE")
_BRD_D = Border(left=_thin_d, right=_thin_d, top=_thin_d, bottom=_thin_d)

_RE_CLIENTE = re.compile(r"Cliente:\s*(.+)")
_RE_CUENTA = re.compile(r"\s*Cuenta:\s*\d+.*")
_RE_NRO_PEDIDO = re.compile(r"NOTA DE PEDIDO\s+Nro\.\s*(\d+)")
_RE_ART_HEADER = re.compile(r"Art[íi]culo\s+Descripci[oó]n", re.IGNORECASE)
_RE_BONIF = re.compile(r"Bonificaciones", re.IGNORECASE)
_RE_ITEM = re.compile(
    r"^\s*(\d+)\s+(.+?)\s+([\d\.]+,\d{4})\s+[\d,]+\s+[\d,\.]+\s+[\d,\.]+\s*$"
)


def darnel_leer_pedido(file_obj) -> list[dict]:
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()

    header_rows = [i for i, row in enumerate(rows) if row[0] == "ID ARTICULO"]
    total_rows = [i for i, row in enumerate(rows) if row[0] and str(row[0]).startswith("TOTAL UM:")]
    if not header_rows or not total_rows:
        raise ValueError("No se encontraron bloques de productos en el pedido.")

    pedido = []
    for header_row, total_row in zip(header_rows, total_rows):
        for row in rows[header_row + 2 : total_row]:
            cod = row[COL_CODIGO_DARNEL]
            if not cod or not str(cod).strip():
                continue
            try:
                cant = int(str(row[COL_CANT_EMPAQUE]).replace(",", "").replace(".", "").strip())
            except (ValueError, TypeError, AttributeError):
                cant = None
            pedido.append({"cod_darnel": str(cod).strip(), "cant_empaque": cant})
    return pedido


def darnel_leer_catalogo(file_obj) -> tuple[dict, dict]:
    df_abm = pd.read_excel(file_obj, sheet_name=HOJA_ABM, header=2)
    df_abm["cd"] = df_abm["Articulo Formularios"].astype(str).str.strip()
    df_abm["cp"] = df_abm["Articulo"].astype(str).str.strip()
    df_abm["nm"] = df_abm["Nombre Articulo"].astype(str).str.strip()
    mapa = (
        df_abm.drop_duplicates("cd")
        .set_index("cd")[["cp", "nm"]]
        .to_dict("index")
    )

    file_obj.seek(0)
    df_res = pd.read_excel(file_obj, sheet_name=HOJA_RESUMEN, header=2)
    df_res["ci"] = df_res["Código"].astype(str).str.strip()
    df_res["zona"] = pd.to_numeric(df_res["Zona"], errors="coerce")
    df_res["cpal"] = pd.to_numeric(df_res["Cant por Pallet"], errors="coerce")
    catalogo = (
        df_res.drop_duplicates("ci")
        .set_index("ci")[["zona", "cpal"]]
        .to_dict("index")
    )
    return mapa, catalogo


def darnel_cruzar(pedido: list[dict], mapa: dict, catalogo: dict) -> tuple[list[dict], list[str]]:
    resultado = []
    sin_match = []
    for item in pedido:
        cod_darnel = item["cod_darnel"]
        abm = mapa.get(cod_darnel)
        if not abm:
            sin_match.append(cod_darnel)
            continue

        cod_pilarica = abm["cp"]
        nombre = abm["nm"]
        cat = catalogo.get(cod_pilarica)
        if not cat:
            sin_match.append(cod_darnel)
            continue

        zona = cat["zona"]
        cant_pal = cat["cpal"]
        cant_emp = item["cant_empaque"]
        pallets = round(cant_emp / cant_pal, 2) if (cant_pal and cant_pal > 0 and cant_emp) else None
        resultado.append(
            {
                "zona": zona,
                "cod_darnel": cod_darnel,
                "cod_pilarica": cod_pilarica,
                "descripcion": nombre,
                "cant_empaque": cant_emp,
                "cant_pallet": int(cant_pal) if cant_pal else None,
                "pallets": pallets,
            }
        )

    resultado.sort(key=lambda row: (row["zona"] or 0, row["cod_darnel"]))
    return resultado, sin_match


def _sd(cell, bold=False, italic=False, bg=None, fg="1F3864", align="left", size=11):
    cell.font = Font(bold=bold, italic=italic, color=fg, name="Arial", size=size)
    cell.fill = PatternFill("solid", start_color=bg) if bg else PatternFill()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _BRD_D


def _set_widths_d(ws):
    for i, width in enumerate(COL_W_D, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _escribir_fila_d(ws, row_index, row_data, item_index):
    bg = "EBF3FB" if item_index % 2 == 0 else "FFFFFF"
    ws.row_dimensions[row_index].height = 17
    _sd(ws.cell(row_index, 1, int(row_data["zona"])), bg=bg, align="center")
    _sd(ws.cell(row_index, 2, row_data["cod_darnel"]), bg=bg)
    _sd(ws.cell(row_index, 3, row_data["cod_pilarica"]), bg=bg)
    _sd(ws.cell(row_index, 4, row_data["descripcion"]), bg=bg)
    c5 = ws.cell(row_index, 5, row_data["cant_empaque"])
    _sd(c5, bg=bg, align="right")
    c5.number_format = "#,##0"
    c6 = ws.cell(row_index, 6, row_data["cant_pallet"])
    _sd(c6, bg=bg, align="right")
    c6.number_format = "#,##0"
    c7 = ws.cell(row_index, 7, row_data["pallets"])
    _sd(c7, bg=bg, align="right")
    c7.number_format = "#,##0.00"


def _escribir_total_d(ws, row_index, first_row, last_row, label="TOTAL"):
    ws.row_dimensions[row_index].height = 19
    for col in range(1, 8):
        _sd(ws.cell(row_index, col, ""), bold=True, bg="DDEBF7")
    _sd(ws.cell(row_index, 4, label), bold=True, bg="DDEBF7", align="right")
    ct = ws.cell(row_index, 5, f"=SUM(E{first_row}:E{last_row})")
    _sd(ct, bold=True, bg="DDEBF7", align="right")
    ct.number_format = "#,##0"
    cp = ws.cell(row_index, 7, f"=SUM(G{first_row}:G{last_row})")
    _sd(cp, bold=True, bg="DDEBF7", align="right")
    cp.number_format = "#,##0.00"


def darnel_generar_excel(resultado: list[dict], nombre_pedido: str) -> bytes:
    wb = Workbook()
    zonas = sorted(set(row["zona"] for row in resultado))

    ws = wb.active
    ws.title = "Resumen General"
    ws.sheet_view.showGridLines = False
    _set_widths_d(ws)

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:G1")
    _sd(ws["A1"], bold=True, bg="1F4E79", fg="FFFFFF", align="center", size=15)
    ws["A1"] = f"PEDIDO {nombre_pedido} · RECEPCIÓN POR ZONA"

    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:G2")
    _sd(ws["A2"], italic=True, bg="2E75B6", fg="FFFFFF", align="center", size=10)
    ws["A2"] = "Agrupado por Zona"

    ws.row_dimensions[3].height = 22
    for col, header in enumerate(HEADERS_D, start=1):
        _sd(ws.cell(3, col, header), bold=True, bg="2E75B6", fg="FFFFFF", align="center")

    row_index = 4
    for zona in zonas:
        zona_int = int(zona)
        items = [row for row in resultado if row["zona"] == zona]
        bgz = ZONE_BG.get(zona_int, "2E75B6")
        fgz = ZONE_FG.get(zona_int, "FFFFFF")

        ws.row_dimensions[row_index].height = 20
        ws.merge_cells(f"A{row_index}:G{row_index}")
        _sd(
            ws.cell(row_index, 1, f" ▶ ZONA {zona_int}"),
            bold=True,
            bg=bgz,
            fg=fgz,
            align="left",
            size=12,
        )
        row_index += 1
        first_row = row_index
        for item_index, row_data in enumerate(items):
            _escribir_fila_d(ws, row_index, row_data, item_index)
            row_index += 1
        _escribir_total_d(ws, row_index, first_row, row_index - 1, f"TOTAL ZONA {zona_int}")
        row_index += 1

    ws.row_dimensions[row_index].height = 24
    for col in range(1, 8):
        _sd(ws.cell(row_index, col, ""), bold=True, bg="1F4E79", fg="FFFFFF")
    total_label = ws.cell(row_index, 4, "TOTAL GENERAL")
    total_label.font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    total_label.border = _BRD_D
    for col, value, number_format in [
        (5, sum(row["cant_empaque"] for row in resultado if row["cant_empaque"]), "#,##0"),
        (
            7,
            round(sum(row["pallets"] for row in resultado if row["pallets"]), 2),
            "#,##0.00",
        ),
    ]:
        cell = ws.cell(row_index, col, value)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _BRD_D
        cell.number_format = number_format

    for zona in zonas:
        zona_int = int(zona)
        items = [row for row in resultado if row["zona"] == zona]
        bgz = ZONE_BG.get(zona_int, "2E75B6")
        fgz = ZONE_FG.get(zona_int, "FFFFFF")
        zone_sheet = wb.create_sheet(f"Zona {zona_int}")
        zone_sheet.sheet_view.showGridLines = False
        _set_widths_d(zone_sheet)
        zone_sheet.row_dimensions[1].height = 34
        zone_sheet.merge_cells("A1:G1")
        _sd(zone_sheet["A1"], bold=True, bg=bgz, fg=fgz, align="center", size=14)
        zone_sheet["A1"] = f"ZONA {zona_int} — PEDIDO {nombre_pedido}"
        for col, header in enumerate(HEADERS_D, start=1):
            _sd(zone_sheet.cell(2, col, header), bold=True, bg=bgz, fg=fgz, align="center")
        for item_index, row_data in enumerate(items):
            _escribir_fila_d(zone_sheet, 3 + item_index, row_data, item_index)
        total_row = 3 + len(items)
        _escribir_total_d(zone_sheet, total_row, 3, total_row - 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def zaplast_load_masterdata() -> pd.DataFrame:
    raw = pd.read_excel(MASTERDATA_PATH, sheet_name="Masterdata", header=None)
    headers = raw.iloc[2].tolist()
    raw.columns = headers
    raw = raw.iloc[3:].reset_index(drop=True)
    raw = raw.rename(columns={raw.columns[0]: "Codigo"})

    col_pallet = next(
        (
            col
            for col in raw.columns
            if "PALLET" in str(col).upper() or "CANTIDAD" in str(col).upper()
        ),
        None,
    )
    if col_pallet is None:
        raise ValueError("No se encontró columna de Cantidad por Pallet en Masterdata")

    raw = raw.rename(columns={col_pallet: "Cantidad por Pallet"})
    masterdata = raw[["Codigo", "ARTICULO", "ZONA", "Cantidad por Pallet"]].copy()
    masterdata["Codigo"] = pd.to_numeric(masterdata["Codigo"], errors="coerce")
    masterdata = masterdata.dropna(subset=["Codigo"])
    masterdata["Codigo"] = masterdata["Codigo"].astype(int)
    masterdata["ZONA"] = pd.to_numeric(masterdata["ZONA"], errors="coerce")
    masterdata["Cantidad por Pallet"] = pd.to_numeric(
        masterdata["Cantidad por Pallet"], errors="coerce"
    )
    masterdata["ARTICULO_NORM"] = masterdata["ARTICULO"].str.upper().str.strip()
    return masterdata


def zaplast_parse_pdf(file_bytes: bytes, filename: str) -> list[dict]:
    rows = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if not text:
                continue

            lines = text.split("\n")

            cliente = ""
            nro_pedido = ""
            for line in lines:
                if not cliente:
                    m = _RE_CLIENTE.search(line)
                    if m:
                        cliente = _RE_CUENTA.sub("", m.group(1)).strip()
                if not nro_pedido:
                    m = _RE_NRO_PEDIDO.search(line)
                    if m:
                        nro_pedido = m.group(1)
                if cliente and nro_pedido:
                    break

            in_items = False
            for line in lines:
                if _RE_ART_HEADER.search(line):
                    in_items = True
                    continue
                if _RE_BONIF.search(line):
                    break
                if in_items:
                    m = _RE_ITEM.match(line)
                    if m:
                        rows.append(
                            {
                                "Nro Pedido": nro_pedido,
                                "Cliente": cliente,
                                "Código Artículo": int(m.group(1)),
                                "Descripción": m.group(2).strip(),
                                "Cantidad": float(
                                    m.group(3).replace(".", "").replace(",", ".")
                                ),
                                "Archivo PDF": filename,
                            }
                        )
    return rows


def zaplast_procesar(pdf_files: list[tuple[str, bytes]], masterdata: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    all_rows = []
    for filename, file_bytes in pdf_files:
        all_rows.extend(zaplast_parse_pdf(file_bytes, filename))

    if not all_rows:
        return pd.DataFrame(), []

    df = pd.DataFrame(all_rows)
    df = df.merge(
        masterdata[["Codigo", "ZONA", "Cantidad por Pallet", "ARTICULO_NORM"]],
        left_on="Código Artículo",
        right_on="Codigo",
        how="left",
    )

    sin_mask = df["ZONA"].isna()
    if sin_mask.any():
        desc_map = (
            masterdata.drop_duplicates("ARTICULO_NORM")
            .set_index("ARTICULO_NORM")[["ZONA", "Cantidad por Pallet"]]
        )
        for idx in df[sin_mask].index:
            key = str(df.at[idx, "Descripción"]).upper().strip()
            if key in desc_map.index:
                df.at[idx, "ZONA"] = desc_map.at[key, "ZONA"]
                df.at[idx, "Cantidad por Pallet"] = desc_map.at[key, "Cantidad por Pallet"]

    sin_match = [
        f"[{row['Código Artículo']}] {row['Descripción']}"
        for _, row in df[df["ZONA"].isna()].iterrows()
    ]
    df["ZONA"] = pd.to_numeric(df["ZONA"], errors="coerce")
    df = df.sort_values(["ZONA", "Cliente", "Código Artículo"]).reset_index(drop=True)
    return df, sin_match


def _thin_z():
    side = Side(style="thin", color="B8CCE4")
    return Border(left=side, right=side, top=side, bottom=side)


def _sz(cell, bold=False, bg=None, fg="000000", size=10, halign="left", num_fmt=None):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.border = _thin_z()
    if num_fmt:
        cell.number_format = num_fmt


def _write_sheet_z(ws, df_all, titulo, single_zona=None):
    ncols = len(COL_HEADERS_Z)
    last_col = get_column_letter(ncols)
    ws.sheet_view.showGridLines = False

    for i, width in enumerate(COL_WIDTHS_Z, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.merge_cells(f"A1:{last_col}1")
    title = ws.cell(1, 1, titulo)
    title.font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last_col}2")
    subtitle = ws.cell(2, 1, "Agrupado por Zona" if single_zona is None else "")
    subtitle.font = Font(name="Arial", italic=True, size=10, color="555555")
    subtitle.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    for col_i, header in enumerate(COL_HEADERS_Z, start=1):
        _sz(
            ws.cell(3, col_i, header),
            bold=True,
            bg=COLOR_HEADER_BG,
            fg=COLOR_HEADER_FG,
            halign="center",
        )
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "A4"

    row_index = 4
    zonas = [single_zona] if single_zona is not None else sorted(df_all["ZONA"].dropna().unique())
    grand_pallets = 0.0
    grand_uds = 0

    for zona_num in zonas:
        zona_int = int(zona_num)
        df_zona = df_all if single_zona is not None else df_all[df_all["ZONA"] == zona_num]

        ws.merge_cells(f"A{row_index}:{last_col}{row_index}")
        _sz(
            ws.cell(row_index, 1, f"▶  ZONA  {zona_int}"),
            bold=True,
            bg=COLOR_ZONA_BG,
            fg=COLOR_ZONA_FG,
        )
        for col_i in range(2, ncols + 1):
            _sz(ws.cell(row_index, col_i), bg=COLOR_ZONA_BG, fg=COLOR_ZONA_FG)
        ws.row_dimensions[row_index].height = 16
        row_index += 1

        zona_pallets = 0.0
        for item_index, (_, row) in enumerate(df_zona.iterrows()):
            bg = COLOR_ROW_ODD if item_index % 2 == 0 else COLOR_ROW_EVEN
            pallets = (
                round(row["Cantidad"] / row["Cantidad por Pallet"], 2)
                if pd.notna(row["Cantidad por Pallet"]) and row["Cantidad por Pallet"]
                else 0.0
            )
            zona_pallets += pallets
            values = [
                zona_int,
                int(row["Código Artículo"]),
                row["Descripción"],
                row["Cliente"],
                int(row["Cantidad"]),
                int(row["Cantidad por Pallet"]) if pd.notna(row["Cantidad por Pallet"]) else "",
                pallets,
            ]
            for col_i, value in enumerate(values, start=1):
                _sz(
                    ws.cell(row_index, col_i, value),
                    bg=bg,
                    halign="left" if col_i in (3, 4) else "center",
                    num_fmt="0.00" if col_i == ncols else None,
                )
            ws.row_dimensions[row_index].height = 15
            row_index += 1

        desc_col = COL_HEADERS_Z.index("Descripción") + 1
        for col_i in range(1, ncols + 1):
            _sz(
                ws.cell(row_index, col_i),
                bold=True,
                bg=COLOR_TOTAL_BG,
                fg=COLOR_TOTAL_FG,
                halign="center",
            )
        ws.cell(row_index, desc_col, f"TOTAL ZONA {zona_int}").font = Font(
            name="Arial", bold=True, size=10, color=COLOR_TOTAL_FG
        )
        ws.cell(row_index, desc_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_index, ncols, round(zona_pallets, 2)).number_format = "0.00"
        ws.cell(row_index, ncols).font = Font(
            name="Arial", bold=True, size=10, color=COLOR_TOTAL_FG
        )
        ws.cell(row_index, ncols).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_index].height = 16
        row_index += 1

        grand_pallets += zona_pallets
        grand_uds += int(df_zona["Cantidad"].sum())

    if single_zona is None:
        cant_col = COL_HEADERS_Z.index("Cantidad") + 1
        desc_col = COL_HEADERS_Z.index("Descripción") + 1
        for col_i in range(1, ncols + 1):
            _sz(
                ws.cell(row_index, col_i),
                bold=True,
                bg=COLOR_GTOTAL_BG,
                fg=COLOR_GTOTAL_FG,
                halign="center",
            )
        ws.cell(row_index, desc_col, "TOTAL GENERAL").font = Font(
            name="Arial", bold=True, size=10, color=COLOR_GTOTAL_FG
        )
        ws.cell(row_index, desc_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_index, cant_col, grand_uds).font = Font(
            name="Arial", bold=True, size=10, color=COLOR_GTOTAL_FG
        )
        ws.cell(row_index, cant_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row_index, ncols, round(grand_pallets, 2)).number_format = "0.00"
        ws.cell(row_index, ncols).font = Font(
            name="Arial", bold=True, size=10, color=COLOR_GTOTAL_FG
        )
        ws.cell(row_index, ncols).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_index].height = 18


def zaplast_generar_excel(df: pd.DataFrame) -> bytes:
    pedidos_label = ", ".join(f"Nro {pedido}" for pedido in sorted(df["Nro Pedido"].unique()))
    titulo_resumen = f"PEDIDOS {pedidos_label}  ·  RECEPCIÓN POR ZONA"

    wb = Workbook()
    resumen_sheet = wb.active
    resumen_sheet.title = "Resumen General"
    _write_sheet_z(resumen_sheet, df, titulo_resumen)

    for zona_num in sorted(df["ZONA"].dropna().unique()):
        zona_int = int(zona_num)
        df_zona = df[df["ZONA"] == zona_num].reset_index(drop=True)
        ws_zona = wb.create_sheet(title=f"Zona {zona_int}")
        _write_sheet_z(
            ws_zona,
            df_zona,
            f"ZONA {zona_int}  —  PEDIDOS {pedidos_label}",
            single_zona=zona_int,
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _safe_int(value):
    if pd.isna(value) or value == "":
        return ""
    return int(value)


def _safe_float(value, digits=2):
    if pd.isna(value) or value == "":
        return ""
    return round(float(value), digits)


def _darnel_row_to_preview(row: dict) -> dict:
    return {
        "Zona": _safe_int(row["zona"]),
        "Código Darnel": row["cod_darnel"],
        "Código Pilarica": row["cod_pilarica"],
        "Descripción": row["descripcion"],
        "Cant x Un Empaque": _safe_int(row["cant_empaque"]) if row["cant_empaque"] else "",
        "Uds/Pallet": _safe_int(row["cant_pallet"]) if row["cant_pallet"] else "",
        "Pallets": _safe_float(row["pallets"]),
    }


def build_darnel_preview(resultado: list[dict], nombre_pedido: str) -> dict:
    zonas = sorted(set(row["zona"] for row in resultado))
    groups = []
    zone_sheets = []

    for zona in zonas:
        zona_int = int(zona)
        items = [row for row in resultado if row["zona"] == zona]
        preview_rows = [_darnel_row_to_preview(row) for row in items]
        total_units = sum(row["cant_empaque"] for row in items if row["cant_empaque"])
        total_pallets = round(sum(row["pallets"] for row in items if row["pallets"]), 2)

        groups.append(
            {
                "label": f"Zona {zona_int}",
                "rows": preview_rows,
                "totals": {
                    "label": f"Total Zona {zona_int}",
                    "Cant x Un Empaque": total_units,
                    "Pallets": total_pallets,
                },
            }
        )
        zone_sheets.append(
            {
                "name": f"Zona {zona_int}",
                "headline": f"ZONA {zona_int} — PEDIDO {nombre_pedido}",
                "subheadline": "",
                "headers": HEADERS_D,
                "groups": [
                    {
                        "label": f"Zona {zona_int}",
                        "rows": preview_rows,
                        "totals": {
                            "label": "Total",
                            "Cant x Un Empaque": total_units,
                            "Pallets": total_pallets,
                        },
                    }
                ],
            }
        )

    return {
        "headers": HEADERS_D,
        "summary": {
            "productos": len(resultado),
            "zonas": len(zonas),
            "palletsTotales": round(sum(row["pallets"] for row in resultado if row["pallets"]), 2),
        },
        "zoneSummary": [
            {
                "zona": int(zona),
                "items": len([row for row in resultado if row["zona"] == zona]),
                "units": sum(
                    row["cant_empaque"] for row in resultado if row["zona"] == zona and row["cant_empaque"]
                ),
                "pallets": round(
                    sum(row["pallets"] for row in resultado if row["zona"] == zona and row["pallets"]),
                    2,
                ),
            }
            for zona in zonas
        ],
        "sheets": [
            {
                "name": "Resumen General",
                "headline": f"PEDIDO {nombre_pedido} · RECEPCIÓN POR ZONA",
                "subheadline": "Agrupado por Zona",
                "headers": HEADERS_D,
                "groups": groups,
                "grandTotal": {
                    "label": "Total General",
                    "Cant x Un Empaque": sum(
                        row["cant_empaque"] for row in resultado if row["cant_empaque"]
                    ),
                    "Pallets": round(sum(row["pallets"] for row in resultado if row["pallets"]), 2),
                },
            },
            *zone_sheets,
        ],
    }


def build_zaplast_preview(df: pd.DataFrame) -> dict:
    zonas = sorted(df["ZONA"].dropna().unique())
    groups = []
    zone_sheets = []

    for zona in zonas:
        zona_int = int(zona)
        df_zona = df[df["ZONA"] == zona].reset_index(drop=True)
        preview_rows = []
        for _, row in df_zona.iterrows():
            pallets = (
                round(row["Cantidad"] / row["Cantidad por Pallet"], 2)
                if pd.notna(row["Cantidad por Pallet"]) and row["Cantidad por Pallet"]
                else 0.0
            )
            preview_rows.append(
                {
                    "Zona": zona_int,
                    "Código Artículo": int(row["Código Artículo"]),
                    "Descripción": row["Descripción"],
                    "Cliente": row["Cliente"],
                    "Cantidad": int(row["Cantidad"]),
                    "Uds/Pallet": int(row["Cantidad por Pallet"])
                    if pd.notna(row["Cantidad por Pallet"])
                    else "",
                    "Pallets": round(pallets, 2),
                }
            )

        zona_pallets = round(sum(row["Pallets"] for row in preview_rows), 2)
        zona_unidades = sum(row["Cantidad"] for row in preview_rows)
        groups.append(
            {
                "label": f"Zona {zona_int}",
                "rows": preview_rows,
                "totals": {
                    "label": f"Total Zona {zona_int}",
                    "Cantidad": zona_unidades,
                    "Pallets": zona_pallets,
                },
            }
        )
        zone_sheets.append(
            {
                "name": f"Zona {zona_int}",
                "headline": f"ZONA {zona_int} — PEDIDOS {', '.join(sorted(df['Nro Pedido'].unique()))}",
                "subheadline": "",
                "headers": COL_HEADERS_Z,
                "groups": [
                    {
                        "label": f"Zona {zona_int}",
                        "rows": preview_rows,
                        "totals": {
                            "label": "Total",
                            "Cantidad": zona_unidades,
                            "Pallets": zona_pallets,
                        },
                    }
                ],
            }
        )

    total_pallets = round(sum(group["totals"]["Pallets"] for group in groups), 2)
    total_unidades = sum(group["totals"]["Cantidad"] for group in groups)
    pedidos_label = ", ".join(f"Nro {pedido}" for pedido in sorted(df["Nro Pedido"].unique()))

    return {
        "headers": COL_HEADERS_Z,
        "summary": {
            "articulos": len(df),
            "zonas": len(zonas),
            "palletsTotales": total_pallets,
        },
        "zoneSummary": [
            {
                "zona": int(zona),
                "items": len(df[df["ZONA"] == zona]),
                "units": int(df[df["ZONA"] == zona]["Cantidad"].sum()),
                "pallets": round(
                    sum(
                        (
                            row["Cantidad"] / row["Cantidad por Pallet"]
                            if pd.notna(row["Cantidad por Pallet"]) and row["Cantidad por Pallet"]
                            else 0.0
                        )
                        for _, row in df[df["ZONA"] == zona].iterrows()
                    ),
                    2,
                ),
            }
            for zona in zonas
        ],
        "sheets": [
            {
                "name": "Resumen General",
                "headline": f"PEDIDOS {pedidos_label} · RECEPCIÓN POR ZONA",
                "subheadline": "Agrupado por Zona",
                "headers": COL_HEADERS_Z,
                "groups": groups,
                "grandTotal": {
                    "label": "Total General",
                    "Cantidad": total_unidades,
                    "Pallets": total_pallets,
                },
            },
            *zone_sheets,
        ],
    }
