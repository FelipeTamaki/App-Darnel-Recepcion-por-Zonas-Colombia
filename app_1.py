"""
app.py — Hub de herramientas de Recepción por Zonas
=====================================================
Funcionalidades:
  1. Recepción por Zonas — Colombia (Darnel/Pilarica)
  2. Notas de Pedido ZAPLAST → Reporte por Zona

Para correr localmente:
    pip install streamlit pandas openpyxl pdfplumber
    streamlit run app.py
"""

import io
import re
from pathlib import Path

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
MASTERDATA_PATH = Path(__file__).parent / "Masterdata.xlsx"

# ── Constantes Darnel (func 1) ────────────────────────────────────────────────
HOJA_ABM     = "ABM ART. DEPURADO 23.02.2025"
HOJA_RESUMEN = "RESUMEN"
COL_CODIGO_DARNEL = 0
COL_CANT_EMPAQUE  = 17

ZONE_BG = {1: "1F4E79", 2: "2E75B6", 3: "0070C0",
           4: "00B0F0", 5: "9DC3E6", 6: "BDD7EE"}
ZONE_FG = {1: "FFFFFF", 2: "FFFFFF", 3: "FFFFFF",
           4: "1F3864", 5: "1F3864", 6: "1F3864"}
HEADERS_D = ["Zona", "Código Darnel", "Código Pilarica", "Descripción",
             "Cant x Un Empaque", "Uds/Pallet", "Pallets"]
COL_W_D   = [8, 18, 17, 38, 18, 13, 13]

# ── Constantes ZAPLAST (func 2) ───────────────────────────────────────────────
COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ZONA_BG   = "2E75B6"
COLOR_ZONA_FG   = "FFFFFF"
COLOR_TOTAL_BG  = "D6E4F0"
COLOR_TOTAL_FG  = "1F4E79"
COLOR_GTOTAL_BG = "1F4E79"
COLOR_GTOTAL_FG = "FFFFFF"
COLOR_ROW_ODD   = "FFFFFF"
COLOR_ROW_EVEN  = "EBF3FB"

COL_HEADERS_Z = ["Zona", "Código Artículo", "Descripción", "Cliente",
                 "Cantidad", "Uds/Pallet", "Pallets"]
COL_WIDTHS_Z  = [8, 18, 35, 30, 12, 12, 10]


# ══════════════════════════════════════════════════════════════════════════════
# CSS COMPARTIDO
# ══════════════════════════════════════════════════════════════════════════════
CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.block-container { max-width: 820px; padding-top: 2rem; }
h1 { font-family: 'IBM Plex Mono', monospace !important; font-size: 1.5rem !important; letter-spacing: -0.5px; }
h3 { font-family: 'IBM Plex Mono', monospace !important; font-size: 1rem !important; color: #2E75B6; }

/* ── Home cards ── */
.home-card {
    border: 2px solid #e0eaf5;
    border-radius: 12px;
    padding: 1.6rem 1.4rem;
    cursor: pointer;
    transition: border-color 0.2s, box-shadow 0.2s;
    background: #f8fbff;
    height: 100%;
}
.home-card:hover { border-color: #2E75B6; box-shadow: 0 4px 16px rgba(30,78,121,0.10); }
.home-card-icon  { font-size: 2.2rem; margin-bottom: 0.5rem; }
.home-card-title {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.95rem;
    font-weight: 600; color: #1F4E79; margin-bottom: 0.4rem;
}
.home-card-desc  { font-size: 0.82rem; color: #555; line-height: 1.5; }

/* ── Misc ── */
.upload-label {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.75rem; font-weight: 600;
    color: #1F4E79; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 0.4rem;
}
.stat-box {
    background: #1F4E79; color: white; border-radius: 8px;
    padding: 1rem 1.2rem; text-align: center; font-family: 'IBM Plex Mono', monospace;
}
.stat-num   { font-size: 2rem; font-weight: 600; line-height: 1; }
.stat-label { font-size: 0.7rem; opacity: 0.7; margin-top: 4px; letter-spacing: 1px; text-transform: uppercase; }

.zone-row {
    display: flex; align-items: center; gap: 12px; padding: 8px 12px;
    border-radius: 6px; margin-bottom: 6px; background: #f8fbff;
    border-left: 4px solid #2E75B6; font-family: 'IBM Plex Mono', monospace; font-size: 0.85rem;
}
.zone-badge {
    background: #1F4E79; color: white; border-radius: 4px; padding: 2px 8px;
    font-weight: 600; font-size: 0.75rem; min-width: 64px; text-align: center;
}
.divider { height: 1px; background: #e0eaf5; margin: 1.4rem 0; }
.warn-box {
    background: #fff8e1; border-left: 4px solid #f0a500;
    padding: 0.8rem 1rem; border-radius: 6px; font-size: 0.85rem; margin-top: 0.5rem;
}
.pdf-tag {
    display: inline-block; background: #e8f4fd; color: #1F4E79; border-radius: 4px;
    padding: 2px 8px; font-family: 'IBM Plex Mono', monospace;
    font-size: 0.72rem; margin: 2px 3px;
}
.back-btn { font-size: 0.8rem; color: #2E75B6; cursor: pointer; }
</style>
"""


# ══════════════════════════════════════════════════════════════════════════════
# LÓGICA — FUNCIONALIDAD 1 (DARNEL)
# ══════════════════════════════════════════════════════════════════════════════

def darnel_leer_pedido(file) -> list:
    wb   = load_workbook(file, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()

    header_rows = [i for i, r in enumerate(rows) if r[0] == "ID ARTICULO"]
    total_rows  = [i for i, r in enumerate(rows)
                   if r[0] and str(r[0]).startswith("TOTAL UM:")]
    if not header_rows or not total_rows:
        raise ValueError("No se encontraron bloques de productos en el pedido.")

    pedido = []
    for h, t in zip(header_rows, total_rows):
        for row in rows[h + 2: t]:
            cod = row[COL_CODIGO_DARNEL]
            if not cod or not str(cod).strip():
                continue
            try:
                cant = int(str(row[COL_CANT_EMPAQUE]).replace(",", "").replace(".", "").strip())
            except (ValueError, TypeError, AttributeError):
                cant = None
            pedido.append({"cod_darnel": str(cod).strip(), "cant_empaque": cant})
    return pedido


def darnel_leer_catalogo(file) -> tuple:
    df_abm = pd.read_excel(file, sheet_name=HOJA_ABM, header=2)
    df_abm["cd"] = df_abm["Articulo Formularios"].astype(str).str.strip()
    df_abm["cp"] = df_abm["Articulo"].astype(str).str.strip()
    df_abm["nm"] = df_abm["Nombre Articulo"].astype(str).str.strip()
    mapa = (df_abm.drop_duplicates("cd")
                  .set_index("cd")[["cp", "nm"]]
                  .to_dict("index"))

    df_res = pd.read_excel(file, sheet_name=HOJA_RESUMEN, header=2)
    df_res["ci"]   = df_res["Código"].astype(str).str.strip()
    df_res["zona"] = pd.to_numeric(df_res["Zona"], errors="coerce")
    df_res["cpal"] = pd.to_numeric(df_res["Cant por Pallet"], errors="coerce")
    catalogo = (df_res.drop_duplicates("ci")
                      .set_index("ci")[["zona", "cpal"]]
                      .to_dict("index"))
    return mapa, catalogo


def darnel_cruzar(pedido, mapa, catalogo) -> tuple:
    resultado, sin_match = [], []
    for p in pedido:
        cd  = p["cod_darnel"]
        abm = mapa.get(cd)
        if not abm:
            sin_match.append(cd)
            continue
        cod_pilarica = abm["cp"]
        nombre       = abm["nm"]
        cat          = catalogo.get(cod_pilarica)
        if not cat:
            sin_match.append(cd)
            continue
        zona     = cat["zona"]
        cant_pal = cat["cpal"]
        cant_emp = p["cant_empaque"]
        pallets  = round(cant_emp / cant_pal, 2) if (cant_pal and cant_pal > 0 and cant_emp) else None
        resultado.append({
            "zona": zona, "cod_darnel": cd, "cod_pilarica": cod_pilarica,
            "descripcion": nombre, "cant_empaque": cant_emp,
            "cant_pallet": int(cant_pal) if cant_pal else None, "pallets": pallets,
        })
    resultado.sort(key=lambda x: (x["zona"] or 0, x["cod_darnel"]))
    return resultado, sin_match


_thin_d = Side(style="thin", color="B0C4DE")
_BRD_D  = Border(left=_thin_d, right=_thin_d, top=_thin_d, bottom=_thin_d)

def _sd(cell, bold=False, italic=False, bg=None, fg="1F3864", align="left", size=11):
    cell.font      = Font(bold=bold, italic=italic, color=fg, name="Arial", size=size)
    cell.fill      = PatternFill("solid", start_color=bg) if bg else PatternFill()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _BRD_D

def _set_widths_d(ws):
    for i, w in enumerate(COL_W_D, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _escribir_fila_d(ws, ri, r, i):
    bg = "EBF3FB" if i % 2 == 0 else "FFFFFF"
    ws.row_dimensions[ri].height = 17
    _sd(ws.cell(ri, 1, int(r["zona"])),          bg=bg, align="center")
    _sd(ws.cell(ri, 2, r["cod_darnel"]),          bg=bg)
    _sd(ws.cell(ri, 3, r["cod_pilarica"]),        bg=bg)
    _sd(ws.cell(ri, 4, r["descripcion"]),         bg=bg)
    c5 = ws.cell(ri, 5, r["cant_empaque"]);   _sd(c5, bg=bg, align="right");  c5.number_format = "#,##0"
    c6 = ws.cell(ri, 6, r["cant_pallet"]);    _sd(c6, bg=bg, align="right");  c6.number_format = "#,##0"
    c7 = ws.cell(ri, 7, r["pallets"]);        _sd(c7, bg=bg, align="right");  c7.number_format = "#,##0.00"

def _escribir_total_d(ws, ri, first, last, label="TOTAL"):
    ws.row_dimensions[ri].height = 19
    for col in range(1, 8):
        _sd(ws.cell(ri, col, ""), bold=True, bg="DDEBF7")
    _sd(ws.cell(ri, 4, label), bold=True, bg="DDEBF7", align="right")
    ct = ws.cell(ri, 5, f"=SUM(E{first}:E{last})")
    _sd(ct, bold=True, bg="DDEBF7", align="right"); ct.number_format = "#,##0"
    cp = ws.cell(ri, 7, f"=SUM(G{first}:G{last})")
    _sd(cp, bold=True, bg="DDEBF7", align="right"); cp.number_format = "#,##0.00"

def darnel_generar_excel(resultado, nombre_pedido) -> bytes:
    import openpyxl
    wb    = openpyxl.Workbook()
    zonas = sorted(set(r["zona"] for r in resultado))

    ws = wb.active
    ws.title = "Resumen General"
    ws.sheet_view.showGridLines = False
    _set_widths_d(ws)

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:G1")
    _sd(ws["A1"], bold=True, bg="1F4E79", fg="FFFFFF", align="center", size=15)
    ws["A1"] = f"PEDIDO {nombre_pedido} · RECEPCIÓN POR ZONA"

    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:G2")
    _sd(ws["A2"], italic=True, bg="2E75B6", fg="FFFFFF", align="center", size=10)
    ws["A2"] = "Agrupado por Zona"

    ws.row_dimensions[3].height = 22
    for col, h in enumerate(HEADERS_D, 1):
        _sd(ws.cell(3, col, h), bold=True, bg="2E75B6", fg="FFFFFF", align="center")

    ri = 4
    for zona in zonas:
        zi    = int(zona)
        items = [r for r in resultado if r["zona"] == zona]
        bgz   = ZONE_BG.get(zi, "2E75B6")
        fgz   = ZONE_FG.get(zi, "FFFFFF")

        ws.row_dimensions[ri].height = 20
        ws.merge_cells(f"A{ri}:G{ri}")
        _sd(ws.cell(ri, 1, f" ▶ ZONA {zi}"), bold=True, bg=bgz, fg=fgz, align="left", size=12)
        ri += 1
        first = ri
        for i, r in enumerate(items):
            _escribir_fila_d(ws, ri, r, i); ri += 1
        _escribir_total_d(ws, ri, first, ri - 1, f"TOTAL ZONA {zi}")
        ri += 1

    ws.row_dimensions[ri].height = 24
    for col in range(1, 8):
        _sd(ws.cell(ri, col, ""), bold=True, bg="1F4E79", fg="FFFFFF")
    c = ws.cell(ri, 4, "TOTAL GENERAL")
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    c.alignment = Alignment(horizontal="right", vertical="center"); c.border = _BRD_D
    for col, val, fmt in [
        (5, sum(r["cant_empaque"] for r in resultado if r["cant_empaque"]), "#,##0"),
        (7, round(sum(r["pallets"] for r in resultado if r["pallets"]), 2), "#,##0.00"),
    ]:
        cx = ws.cell(ri, col, val)
        cx.font      = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        cx.alignment = Alignment(horizontal="right", vertical="center")
        cx.border    = _BRD_D; cx.number_format = fmt

    for zona in zonas:
        zi    = int(zona)
        items = [r for r in resultado if r["zona"] == zona]
        bgz   = ZONE_BG.get(zi, "2E75B6")
        fgz   = ZONE_FG.get(zi, "FFFFFF")
        wz    = wb.create_sheet(f"Zona {zi}")
        wz.sheet_view.showGridLines = False
        _set_widths_d(wz)
        wz.row_dimensions[1].height = 34
        wz.merge_cells("A1:G1")
        _sd(wz["A1"], bold=True, bg=bgz, fg=fgz, align="center", size=14)
        wz["A1"] = f"ZONA {zi} — PEDIDO {nombre_pedido}"
        for col, h in enumerate(HEADERS_D, 1):
            _sd(wz.cell(2, col, h), bold=True, bg=bgz, fg=fgz, align="center")
        for i, r in enumerate(items):
            _escribir_fila_d(wz, 3 + i, r, i)
        tr = 3 + len(items)
        _escribir_total_d(wz, tr, 3, tr - 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# LÓGICA — FUNCIONALIDAD 2 (ZAPLAST)
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def zaplast_load_masterdata() -> pd.DataFrame:
    raw     = pd.read_excel(MASTERDATA_PATH, sheet_name="Masterdata", header=None)
    headers = raw.iloc[2].tolist()
    raw.columns = headers
    raw     = raw.iloc[3:].reset_index(drop=True)
    raw     = raw.rename(columns={raw.columns[0]: "Codigo"})

    col_pallet = next(
        (c for c in raw.columns if "PALLET" in str(c).upper() or "CANTIDAD" in str(c).upper()), None
    )
    if col_pallet is None:
        raise ValueError("No se encontró columna de Cantidad por Pallet en Masterdata")
    raw = raw.rename(columns={col_pallet: "Cantidad por Pallet"})

    md = raw[["Codigo", "ARTICULO", "ZONA", "Cantidad por Pallet"]].copy()
    md["Codigo"]              = pd.to_numeric(md["Codigo"], errors="coerce")
    md                        = md.dropna(subset=["Codigo"])
    md["Codigo"]              = md["Codigo"].astype(int)
    md["ZONA"]                = pd.to_numeric(md["ZONA"], errors="coerce")
    md["Cantidad por Pallet"] = pd.to_numeric(md["Cantidad por Pallet"], errors="coerce")
    md["ARTICULO_NORM"]       = md["ARTICULO"].str.upper().str.strip()
    return md


def zaplast_parse_pdf(file_bytes: bytes, filename: str) -> list[dict]:
    rows = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            lines = text.split("\n")

            cliente = ""
            for line in lines:
                m = re.search(r"Cliente:\s*(.+)", line)
                if m:
                    cliente = re.sub(r"\s*Cuenta:\s*\d+.*", "", m.group(1)).strip()
                    break

            nro_pedido = ""
            for line in lines:
                m = re.search(r"NOTA DE PEDIDO\s+Nro\.\s*(\d+)", line)
                if m:
                    nro_pedido = m.group(1)
                    break

            in_items = False
            for line in lines:
                if re.search(r"Artículo\s+Descripción", line, re.IGNORECASE):
                    in_items = True; continue
                if re.search(r"Bonificaciones", line, re.IGNORECASE):
                    in_items = False; continue
                if in_items:
                    m = re.match(
                        r"^\s*(\d+)\s+(.+?)\s+([\d\.]+,\d{4})\s+[\d,]+\s+[\d,\.]+\s+[\d,\.]+\s*$",
                        line,
                    )
                    if m:
                        rows.append({
                            "Nro Pedido":      nro_pedido,
                            "Cliente":         cliente,
                            "Código Artículo": int(m.group(1)),
                            "Descripción":     m.group(2).strip(),
                            "Cantidad":        float(m.group(3).replace(".", "").replace(",", ".")),
                            "Archivo PDF":     filename,
                        })
    return rows


def zaplast_procesar(pdf_files: list, md: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    all_rows = []
    for f in pdf_files:
        all_rows.extend(zaplast_parse_pdf(f.read(), f.name))

    if not all_rows:
        return pd.DataFrame(), []

    df = pd.DataFrame(all_rows)
    df = df.merge(
        md[["Codigo", "ZONA", "Cantidad por Pallet", "ARTICULO_NORM"]],
        left_on="Código Artículo", right_on="Codigo", how="left"
    )

    sin_mask = df["ZONA"].isna()
    if sin_mask.any():
        desc_map = (md.drop_duplicates("ARTICULO_NORM")
                      .set_index("ARTICULO_NORM")[["ZONA", "Cantidad por Pallet"]])
        for idx in df[sin_mask].index:
            key = str(df.at[idx, "Descripción"]).upper().strip()
            if key in desc_map.index:
                df.at[idx, "ZONA"]               = desc_map.at[key, "ZONA"]
                df.at[idx, "Cantidad por Pallet"] = desc_map.at[key, "Cantidad por Pallet"]

    sin_match = [
        f"[{row['Código Artículo']}] {row['Descripción']}"
        for _, row in df[df["ZONA"].isna()].iterrows()
    ]
    df["ZONA"] = pd.to_numeric(df["ZONA"], errors="coerce")
    df = df.sort_values(["ZONA", "Cliente", "Código Artículo"]).reset_index(drop=True)
    return df, sin_match


def _thin_z():
    s = Side(style="thin", color="B8CCE4")
    return Border(left=s, right=s, top=s, bottom=s)

def _sz(cell, bold=False, bg=None, fg="000000", size=10,
        halign="left", num_fmt=None):
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    if bg:
        cell.fill  = PatternFill("solid", start_color=bg)
    cell.border    = _thin_z()
    if num_fmt:
        cell.number_format = num_fmt

def _write_sheet_z(ws, df_all, titulo, single_zona=None):
    ncols    = len(COL_HEADERS_Z)
    last_col = get_column_letter(ncols)
    ws.sheet_view.showGridLines = False

    for i, w in enumerate(COL_WIDTHS_Z, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f"A1:{last_col}1")
    t = ws.cell(1, 1, titulo)
    t.font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last_col}2")
    s = ws.cell(2, 1, "Agrupado por Zona" if single_zona is None else "")
    s.font      = Font(name="Arial", italic=True, size=10, color="555555")
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    for col_i, h in enumerate(COL_HEADERS_Z, 1):
        _sz(ws.cell(3, col_i, h), bold=True, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG, halign="center")
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "A4"

    r             = 4
    zonas         = [single_zona] if single_zona is not None else sorted(df_all["ZONA"].dropna().unique())
    grand_pallets = 0.0
    grand_uds     = 0

    for zona_num in zonas:
        zona_int = int(zona_num)
        df_zona  = df_all if single_zona is not None else df_all[df_all["ZONA"] == zona_num]

        ws.merge_cells(f"A{r}:{last_col}{r}")
        _sz(ws.cell(r, 1, f"▶  ZONA  {zona_int}"),
            bold=True, bg=COLOR_ZONA_BG, fg=COLOR_ZONA_FG)
        for col_i in range(2, ncols + 1):
            _sz(ws.cell(r, col_i), bg=COLOR_ZONA_BG, fg=COLOR_ZONA_FG)
        ws.row_dimensions[r].height = 16
        r += 1

        zona_pallets = 0.0
        for i, (_, row) in enumerate(df_zona.iterrows()):
            bg      = COLOR_ROW_ODD if i % 2 == 0 else COLOR_ROW_EVEN
            pallets = round(row["Cantidad"] / row["Cantidad por Pallet"], 2) \
                      if pd.notna(row["Cantidad por Pallet"]) and row["Cantidad por Pallet"] else 0.0
            zona_pallets += pallets
            values = [
                zona_int, int(row["Código Artículo"]), row["Descripción"],
                row["Cliente"], int(row["Cantidad"]),
                int(row["Cantidad por Pallet"]) if pd.notna(row["Cantidad por Pallet"]) else "",
                pallets,
            ]
            for col_i, val in enumerate(values, 1):
                _sz(ws.cell(r, col_i, val),
                    bg=bg,
                    halign="left" if col_i in (3, 4) else "center",
                    num_fmt="0.00" if col_i == ncols else None)
            ws.row_dimensions[r].height = 15
            r += 1

        desc_col = COL_HEADERS_Z.index("Descripción") + 1
        for col_i in range(1, ncols + 1):
            _sz(ws.cell(r, col_i), bold=True, bg=COLOR_TOTAL_BG, fg=COLOR_TOTAL_FG, halign="center")
        ws.cell(r, desc_col, f"TOTAL ZONA {zona_int}").font = \
            Font(name="Arial", bold=True, size=10, color=COLOR_TOTAL_FG)
        ws.cell(r, desc_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, ncols, round(zona_pallets, 2)).number_format = "0.00"
        ws.cell(r, ncols).font      = Font(name="Arial", bold=True, size=10, color=COLOR_TOTAL_FG)
        ws.cell(r, ncols).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 16
        r += 1

        grand_pallets += zona_pallets
        grand_uds     += int(df_zona["Cantidad"].sum())

    if single_zona is None:
        cant_col = COL_HEADERS_Z.index("Cantidad") + 1
        desc_col = COL_HEADERS_Z.index("Descripción") + 1
        for col_i in range(1, ncols + 1):
            _sz(ws.cell(r, col_i), bold=True, bg=COLOR_GTOTAL_BG, fg=COLOR_GTOTAL_FG, halign="center")
        ws.cell(r, desc_col, "TOTAL GENERAL").font = \
            Font(name="Arial", bold=True, size=10, color=COLOR_GTOTAL_FG)
        ws.cell(r, desc_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, cant_col, grand_uds).font = \
            Font(name="Arial", bold=True, size=10, color=COLOR_GTOTAL_FG)
        ws.cell(r, cant_col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, ncols, round(grand_pallets, 2)).number_format = "0.00"
        ws.cell(r, ncols).font      = Font(name="Arial", bold=True, size=10, color=COLOR_GTOTAL_FG)
        ws.cell(r, ncols).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18


def zaplast_generar_excel(df: pd.DataFrame) -> bytes:
    pedidos_label = ", ".join(f"Nro {p}" for p in sorted(df["Nro Pedido"].unique()))
    titulo_res    = f"PEDIDOS {pedidos_label}  ·  RECEPCIÓN POR ZONA"

    wb            = Workbook()
    ws_res        = wb.active
    ws_res.title  = "Resumen General"
    _write_sheet_z(ws_res, df, titulo_res)

    for zona_num in sorted(df["ZONA"].dropna().unique()):
        zona_int = int(zona_num)
        df_zona  = df[df["ZONA"] == zona_num].reset_index(drop=True)
        ws_z     = wb.create_sheet(title=f"Zona {zona_int}")
        _write_sheet_z(ws_z, df_zona,
                       f"ZONA {zona_int}  —  PEDIDOS {pedidos_label}",
                       single_zona=zona_int)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS UI COMPARTIDOS
# ══════════════════════════════════════════════════════════════════════════════

def stat_box(value, label):
    st.markdown(
        f'<div class="stat-box"><div class="stat-num">{value}</div>'
        f'<div class="stat-label">{label}</div></div>',
        unsafe_allow_html=True,
    )

def zone_row(zona_int, n_items, uds, pallets):
    st.markdown(
        f'<div class="zone-row">'
        f'<span class="zone-badge">ZONA {zona_int}</span>'
        f'<span>{n_items} artículo{"s" if n_items != 1 else ""}</span>'
        f'<span style="margin-left:auto;opacity:.6">{uds:,} uds · {pallets} pallets</span>'
        f'</div>',
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
# APP PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Recepción por Zonas",
    page_icon="📦",
    layout="centered",
)
st.markdown(CSS, unsafe_allow_html=True)

# ── Estado de navegación ──────────────────────────────────────────────────────
if "pantalla" not in st.session_state:
    st.session_state.pantalla = "home"


# ════════════════════════════════════════════════════════════
# PANTALLA HOME
# ════════════════════════════════════════════════════════════
if st.session_state.pantalla == "home":

    st.markdown("# 📦 Recepción por Zonas")
    st.markdown("Seleccioná la herramienta que necesitás usar.")
    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.markdown("""
        <div class="home-card">
            <div class="home-card-icon">🌎</div>
            <div class="home-card-title">Pedidos Darnel — Colombia</div>
            <div class="home-card-desc">
                Cruzá un pedido Darnel (.xlsx) con el catálogo Pilarica
                y generá el reporte Excel agrupado por zona.
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Ir a Darnel →", use_container_width=True, key="btn_darnel"):
            st.session_state.pantalla = "darnel"
            st.rerun()

    with col2:
        st.markdown("""
        <div class="home-card">
            <div class="home-card-icon">🗂️</div>
            <div class="home-card-title">Notas de Pedido ZAPLAST</div>
            <div class="home-card-desc">
                Subí uno o más PDFs de Notas de Pedido ZAPLAST y
                descargá el reporte por zona con pallets calculados.
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Ir a ZAPLAST →", use_container_width=True, key="btn_zaplast"):
            st.session_state.pantalla = "zaplast"
            st.rerun()


# ════════════════════════════════════════════════════════════
# PANTALLA DARNEL
# ════════════════════════════════════════════════════════════
elif st.session_state.pantalla == "darnel":

    if st.button("← Volver al inicio", key="back_darnel"):
        st.session_state.pantalla = "home"
        st.rerun()

    st.markdown("# 🌎 Pedidos Darnel — Colombia")
    st.markdown("Cruzá el pedido con el catálogo y descargá el reporte agrupado por zona.")
    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<div class="upload-label">① Pedido Darnel (.xlsx)</div>', unsafe_allow_html=True)
        pedido_file = st.file_uploader("", type=["xlsx"], key="pedido_d", label_visibility="collapsed")
        if pedido_file:
            st.success(f"✓ {pedido_file.name}")

    with col2:
        st.markdown('<div class="upload-label">② Catálogo Pilarica (.xlsx)</div>', unsafe_allow_html=True)
        catalogo_file = st.file_uploader("", type=["xlsx"], key="catalogo_d", label_visibility="collapsed")
        if catalogo_file:
            st.success(f"✓ {catalogo_file.name}")

    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

    if pedido_file and catalogo_file:
        if st.button("⚡ Procesar", use_container_width=True, type="primary", key="proc_darnel"):
            with st.spinner("Procesando..."):
                try:
                    pedido             = darnel_leer_pedido(pedido_file)
                    mapa, catalogo     = darnel_leer_catalogo(catalogo_file)
                    resultado, sin_match = darnel_cruzar(pedido, mapa, catalogo)

                    zonas      = sorted(set(r["zona"] for r in resultado))
                    total_u    = sum(r["cant_empaque"] for r in resultado if r["cant_empaque"])
                    total_pal  = round(sum(r["pallets"] for r in resultado if r["pallets"]), 2)
                    nombre_ped = Path(pedido_file.name).stem

                    c1, c2, c3 = st.columns(3)
                    with c1: stat_box(len(resultado), "Productos")
                    with c2: stat_box(len(zonas), "Zonas")
                    with c3: stat_box(f"{total_pal:,.1f}", "Pallets totales")

                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown("### Resumen por zona")
                    for zona in zonas:
                        items   = [r for r in resultado if r["zona"] == zona]
                        zona_u  = sum(r["cant_empaque"] for r in items if r["cant_empaque"])
                        zona_p  = round(sum(r["pallets"] for r in items if r["pallets"]), 2)
                        zone_row(int(zona), len(items), zona_u, zona_p)

                    if sin_match:
                        st.markdown(
                            f'<div class="warn-box">⚠️ <b>{len(sin_match)} producto(s)</b> '
                            f'no encontrados: {", ".join(sin_match)}</div>',
                            unsafe_allow_html=True,
                        )

                    st.markdown("<br>", unsafe_allow_html=True)
                    st.download_button(
                        label="📥 Descargar Excel",
                        data=darnel_generar_excel(resultado, nombre_ped),
                        file_name=f"reporte_por_zona_{nombre_ped}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                except Exception as e:
                    st.error(f"❌ Error al procesar: {e}")
                    st.exception(e)
    else:
        st.info("Subí los dos archivos para habilitar el procesamiento.")


# ════════════════════════════════════════════════════════════
# PANTALLA ZAPLAST
# ════════════════════════════════════════════════════════════
elif st.session_state.pantalla == "zaplast":

    if st.button("← Volver al inicio", key="back_zaplast"):
        st.session_state.pantalla = "home"
        st.rerun()

    st.markdown("# 🗂️ Notas de Pedido ZAPLAST")
    st.markdown("Subí uno o más PDFs y descargá el reporte agrupado por zona.")
    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

    st.markdown('<div class="upload-label">Notas de Pedido (PDF) — podés subir varios a la vez</div>',
                unsafe_allow_html=True)
    pdf_files = st.file_uploader(
        "", type=["pdf"], accept_multiple_files=True,
        key="zaplast_pdfs", label_visibility="collapsed",
    )

    if pdf_files:
        tags = " ".join(f'<span class="pdf-tag">📄 {f.name}</span>' for f in pdf_files)
        st.markdown(tags, unsafe_allow_html=True)

    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

    if pdf_files:
        if st.button("⚡ Procesar", use_container_width=True, type="primary", key="proc_zaplast"):
            with st.spinner("Procesando PDFs..."):
                try:
                    md            = zaplast_load_masterdata()
                    df, sin_match = zaplast_procesar(pdf_files, md)

                    if df.empty:
                        st.error("❌ No se encontraron artículos en los PDFs subidos.")
                        st.stop()

                    zonas         = sorted(df["ZONA"].dropna().unique())
                    total_pallets = round(sum(
                        row["Cantidad"] / row["Cantidad por Pallet"]
                        for _, row in df.iterrows()
                        if pd.notna(row["Cantidad por Pallet"]) and row["Cantidad por Pallet"]
                    ), 2)

                    c1, c2, c3 = st.columns(3)
                    with c1: stat_box(len(df), "Artículos")
                    with c2: stat_box(len(zonas), "Zonas")
                    with c3: stat_box(f"{total_pallets:,.1f}", "Pallets totales")

                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown("### Resumen por zona")
                    for zona_num in zonas:
                        zona_int = int(zona_num)
                        df_zona  = df[df["ZONA"] == zona_num]
                        zona_uds = int(df_zona["Cantidad"].sum())
                        zona_pal = round(sum(
                            r["Cantidad"] / r["Cantidad por Pallet"]
                            for _, r in df_zona.iterrows()
                            if pd.notna(r["Cantidad por Pallet"]) and r["Cantidad por Pallet"]
                        ), 2)
                        zone_row(zona_int, len(df_zona), zona_uds, zona_pal)

                    if sin_match:
                        st.markdown(
                            f'<div class="warn-box">⚠️ <b>{len(sin_match)} artículo(s)</b> '
                            f'no encontrados en el Masterdata: {", ".join(sin_match)}</div>',
                            unsafe_allow_html=True,
                        )

                    st.markdown("<br>", unsafe_allow_html=True)
                    pedidos_label = "_".join(sorted(df["Nro Pedido"].unique()))
                    st.download_button(
                        label="📥 Descargar Excel",
                        data=zaplast_generar_excel(df),
                        file_name=f"pedidos_por_zona_{pedidos_label}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                except Exception as e:
                    st.error(f"❌ Error al procesar: {e}")
                    st.exception(e)
    else:
        st.info("Subí uno o más PDFs de Notas de Pedido ZAPLAST para habilitar el procesamiento.")
